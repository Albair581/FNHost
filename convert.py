from openpyxl import load_workbook
import os
import json

# Load workbook
wb = load_workbook("FN Research.xlsx")

# Load existing JSON (or start fresh)
if os.path.exists("fn-research.json"):
    with open("fn-research.json", "r", encoding="utf-8") as f:
        fine = json.load(f)
else:
    fine = []

# Iterate through every sheet = every org
for sheet in wb.worksheets:
    ORG_NAME = sheet.title

    # Find or create org
    org_obj = next(
        (org for org in fine if org["org"] == ORG_NAME),
        None
    )

    if org_obj is None:
        org_obj = {
            "org": ORG_NAME,
            "branches": []
        }
        fine.append(org_obj)

    # Optional: clear old branches if this is a full re-import
    org_obj["branches"].clear()

    # Loop through rows
    for row in sheet.iter_rows(min_row=2, values_only=True, max_col=13):
        if not row[0]:  # skip empty rows
            continue

        data = {
            "name": row[0] + f"（{row[1]}）",
            "address": row[2],
            "hours": row[6],
            "phone": row[4],
            "fax": row[5],
            "coordinates": [
                float(row[3].split(",")[0]),
                float(row[3].split(",")[1])
            ],
            "transportation": row[7],
            "parking": row[8],
            "accepted_items": row[9],
            "notes": row[10],
            "maps_url": row[11],
            "update_time": str(row[12]),
        }

        org_obj["branches"].append(data)

# Dump back to JSON
with open("fn-research.json", "w", encoding="utf-8") as f:
    json.dump(fine, f, ensure_ascii=False, indent=2)